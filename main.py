import openpyxl

theFile = openpyxl.load_workbook('test.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))

open('readme.txt', 'w').close()

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]

    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEF":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            cell_name_plus_1 = "{}{}".format("B", row)
            if str(currentSheet[cell_name].value).find("Test")!=-1:
                #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name_plus_1].value))
                with open('readme.txt', 'a') as f:
                    f.write("{} starts in line {}".format(currentSheet[cell_name].value,str(currentSheet[cell_name_plus_1].value)) + "\n")
